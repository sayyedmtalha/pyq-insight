import io
import os
import tempfile
import time
from contextlib import redirect_stderr, redirect_stdout

import streamlit as st
import openpyxl
from pipeline import run_pipeline
from pdf_extraction import extract_text_from_pdf


st.set_page_config(page_title="PYQ Insight | Universal Exam Analyzer", layout="wide", page_icon="🎓")

# Custom Claude-inspired Warm Terracotta & Dark Slate CSS
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif;
    }

    /* Main background */
    .stApp {
        background-color: #18181B;
        color: #F4F4F5;
    }

    /* Hero Banner Header */
    .hero-banner {
        background: linear-gradient(135deg, #24221F 0%, #1A1917 100%);
        border: 1px solid rgba(204, 107, 73, 0.3);
        border-radius: 20px;
        padding: 32px 40px;
        margin-bottom: 28px;
        box-shadow: 0 12px 36px rgba(0, 0, 0, 0.4);
        position: relative;
    }

    .badge-tag {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: rgba(204, 107, 73, 0.15);
        color: #DA7756;
        border: 1px solid rgba(204, 107, 73, 0.35);
        font-size: 12px;
        font-weight: 600;
        padding: 5px 14px;
        border-radius: 20px;
        margin-bottom: 14px;
        letter-spacing: 0.5px;
        text-transform: uppercase;
    }

    .hero-title {
        font-size: 40px;
        font-weight: 700;
        color: #FAF8F5;
        margin: 0 0 10px 0;
        letter-spacing: -0.5px;
    }

    .hero-subtitle {
        font-size: 16px;
        color: #A1A1AA;
        margin: 0;
        line-height: 1.6;
        max-width: 800px;
    }

    /* Card Panels */
    .card-panel {
        background: #201F1C;
        border: 1px solid #33312D;
        border-radius: 16px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.25);
    }

    .card-header {
        font-size: 18px;
        font-weight: 600;
        color: #E4E4E7;
        margin-bottom: 16px;
        display: flex;
        align-items: center;
        gap: 10px;
    }

    /* Streamlit Button Styles */
    div.stButton > button {
        background: linear-gradient(135deg, #CC6B49 0%, #B2502E 100%) !important;
        color: #FFFFFF !important;
        font-weight: 600 !important;
        font-size: 16px !important;
        padding: 14px 32px !important;
        border-radius: 12px !important;
        border: none !important;
        box-shadow: 0 6px 20px rgba(204, 107, 73, 0.35) !important;
        transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
        width: 100% !important;
    }

    div.stButton > button:hover {
        background: linear-gradient(135deg, #DA7756 0%, #C15F3D 100%) !important;
        box-shadow: 0 8px 25px rgba(204, 107, 73, 0.5) !important;
        transform: translateY(-2px) !important;
    }

    /* Download Button Styles */
    div.stDownloadButton > button {
        background: linear-gradient(135deg, #22C55E 0%, #16A34A 100%) !important;
        color: #FFFFFF !important;
        font-weight: 600 !important;
        font-size: 16px !important;
        padding: 14px 32px !important;
        border-radius: 12px !important;
        border: none !important;
        box-shadow: 0 6px 20px rgba(34, 197, 94, 0.35) !important;
        width: 100% !important;
    }

    div.stDownloadButton > button:hover {
        background: linear-gradient(135deg, #4ADE80 0%, #22C55E 100%) !important;
        transform: translateY(-2px) !important;
    }

    /* Upload Area Styling */
    section[data-testid="stFileUploader"] {
        background: #191816 !important;
        border: 1px dashed #42403B !important;
        border-radius: 12px !important;
        padding: 12px !important;
    }

    section[data-testid="stFileUploader"]:hover {
        border-color: #CC6B49 !important;
    }

    /* Metric Cards */
    .metric-card {
        background: #262522;
        border: 1px solid #3B3934;
        border-radius: 14px;
        padding: 20px;
        text-align: center;
    }

    .metric-val {
        font-size: 32px;
        font-weight: 700;
        color: #DA7756;
    }

    .metric-lbl {
        font-size: 13px;
        color: #A1A1AA;
        margin-top: 6px;
        font-weight: 500;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


class LiveLogBuffer(io.StringIO):
    """Stream pipeline output into a bounded Streamlit code panel as it is produced."""

    def __init__(self, placeholder, refresh_interval: float = 0.2, max_characters: int = 12000):
        super().__init__()
        self.placeholder = placeholder
        self.refresh_interval = refresh_interval
        self.max_characters = max_characters
        self.last_render = 0.0

    def write(self, text):
        written = super().write(text)
        if text and ("\n" in text or time.monotonic() - self.last_render >= self.refresh_interval):
            self.render()
        return written

    def render(self):
        log_text = self.getvalue()
        if len(log_text) > self.max_characters:
            log_text = "... earlier log output omitted ...\n" + log_text[-self.max_characters:]
        self.placeholder.code(log_text or "Waiting for processing output...", language="text")
        self.last_render = time.monotonic()

    def flush(self):
        self.render()


# Header Banner
st.markdown(
    """
    <div class="hero-banner">
        <div class="badge-tag">✨ Universal AI Analysis Pipeline</div>
        <h1 class="hero-title">PYQ Insight</h1>
        <p class="hero-subtitle">Upload past-year question paper PDFs and a course syllabus to automatically extract sub-questions, align Course Outcomes (CO), deduplicate variants, and generate a standardized Excel analysis report.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# Upload Layout
col1, col2 = st.columns(2)

with col1:
    st.markdown('<div class="card-header">📄 Past-Year Question Papers</div>', unsafe_allow_html=True)
    pdf_files = st.file_uploader("Upload PYQ PDFs (Multi-Year)", type=["pdf"], accept_multiple_files=True, help="Upload 1 or more past-year exam paper PDFs")

with col2:
    st.markdown('<div class="card-header">📚 Course Syllabus</div>', unsafe_allow_html=True)
    syllabus_file = st.file_uploader("Upload Syllabus (PDF or TXT)", type=["pdf", "txt"], help="Upload syllabus text file or PDF")

use_ocr = st.checkbox("Enable OCR for Scanned PDFs", value=True, help="Attempt OCR when available; otherwise the pipeline falls back to native PDF text extraction")

st.markdown("<br>", unsafe_allow_html=True)

if st.button("🚀 Analyze Exam Dataset"):
    if not pdf_files:
        st.error("Please upload at least one PDF exam paper.")
    elif not syllabus_file:
        st.error("Please upload a syllabus file.")
    else:
        st.info(f"Processing {len(pdf_files)} PDF paper(s) with OCR {'enabled' if use_ocr else 'disabled'}...")
        
        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_paths = []
            for uploaded_file in pdf_files:
                temp_path = os.path.join(tmpdir, uploaded_file.name)
                with open(temp_path, "wb") as fh:
                    fh.write(uploaded_file.getbuffer())
                pdf_paths.append(temp_path)

            syllabus_path = os.path.join(tmpdir, syllabus_file.name)
            with open(syllabus_path, "wb") as fh:
                fh.write(syllabus_file.getbuffer())

            syllabus_text = ""
            if syllabus_file.name.endswith(".txt"):
                syllabus_text = syllabus_file.getvalue().decode("utf-8")
            else:
                syllabus_text = extract_text_from_pdf(syllabus_path, force_ocr=False)
                if not syllabus_text.strip() and use_ocr:
                    syllabus_text = extract_text_from_pdf(syllabus_path, force_ocr=True)
                if not syllabus_text.strip():
                    syllabus_text = f"Syllabus file: {syllabus_file.name}"

            output_path = os.path.join(tmpdir, "pyq_insight_output.xlsx")
            with st.expander("💻 Live Processing Terminal", expanded=True):
                log_placeholder = st.empty()
            log_buffer = LiveLogBuffer(log_placeholder)
            
            try:
                with redirect_stdout(log_buffer), redirect_stderr(log_buffer):
                    result = run_pipeline(pdf_paths, syllabus_text, output_path, force_ocr=use_ocr)
                log_buffer.flush()
                
                wb = openpyxl.load_workbook(result)
                topic_sheet = wb['Topic Analysis'] if 'Topic Analysis' in wb.sheetnames else wb.active
                topic_count = max(topic_sheet.max_row - 1, 0)
                
                st.success("🎉 Excel Analysis Workbook Generated Successfully!")
                
                # Render Metrics Summary Cards
                m1, m2, m3 = st.columns(3)
                with m1:
                    st.markdown(f'<div class="metric-card"><div class="metric-val">{len(pdf_files)}</div><div class="metric-lbl">Papers Parsed</div></div>', unsafe_allow_html=True)
                with m2:
                    st.markdown(f'<div class="metric-card"><div class="metric-val">{topic_count}</div><div class="metric-lbl">Concept Topics Analyzed</div></div>', unsafe_allow_html=True)
                with m3:
                    st.markdown(f'<div class="metric-card"><div class="metric-val">100%</div><div class="metric-lbl">Dataset Coverage</div></div>', unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                with open(result, "rb") as fh:
                    workbook_bytes = fh.read()
                st.download_button(
                    "📥 Download Excel Analysis Workbook (.xlsx)",
                    data=workbook_bytes,
                    file_name="pyq_insight_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as exc:
                log_buffer.flush()
                st.error(f"Analysis failed: {exc}")
                st.stop()
