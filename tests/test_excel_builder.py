import os
import tempfile
import unittest
from openpyxl import load_workbook

from excel_builder import build_workbook


class WorkbookBuilderTests(unittest.TestCase):
    def test_build_workbook_uses_topic_analysis_structure(self):
        raw_records = [
            {
                "question_number": "1(a)",
                "text": "Explain the Newton-Raphson method.",
                "topic": "Unit I - Numerical Methods",
                "subtopic": "Root finding",
                "difficulty": "Medium",
                "question_type": "Derivation",
                "confidence": 0.82,
                "manual_review": False,
                "year": 2021,
                "marks": 5,
                "source_page": 2,
            },
            {
                "question_number": "3(b)",
                "text": "Derive the formula.",
                "topic": "Unit I - Numerical Methods",
                "subtopic": "Root finding",
                "difficulty": "Hard",
                "question_type": "Derivation",
                "confidence": 0.75,
                "manual_review": False,
                "year": 2021,
                "marks": 8,
                "source_page": 7,
            },
            {
                "question_number": "2(a)",
                "text": "State the theorem.",
                "topic": "Unit II - Probability",
                "subtopic": "Random variables",
                "difficulty": "Easy",
                "question_type": "Theoretical",
                "confidence": 0.66,
                "manual_review": True,
                "year": 2022,
                "marks": 3,
                "source_page": 5,
            },
        ]
import os
import tempfile
import unittest
from openpyxl import load_workbook

from excel_builder import build_workbook


class WorkbookBuilderTests(unittest.TestCase):
    def test_build_workbook_uses_topic_analysis_structure(self):
        raw_records = [
            {
                "question_number": "1(a)",
                "text": "Explain the Newton-Raphson method.",
                "topic": "Unit I - Numerical Methods",
                "subtopic": "Root finding",
                "difficulty": "Medium",
                "question_type": "Derivation",
                "confidence": 0.82,
                "manual_review": False,
                "year": 2021,
                "marks": 5,
                "source_page": 2,
            },
            {
                "question_number": "3(b)",
                "text": "Derive the formula.",
                "topic": "Unit I - Numerical Methods",
                "subtopic": "Root finding",
                "difficulty": "Hard",
                "question_type": "Derivation",
                "confidence": 0.75,
                "manual_review": False,
                "year": 2021,
                "marks": 8,
                "source_page": 7,
            },
            {
                "question_number": "2(a)",
                "text": "State the theorem.",
                "topic": "Unit II - Probability",
                "subtopic": "Random variables",
                "difficulty": "Easy",
                "question_type": "Theoretical",
                "confidence": 0.66,
                "manual_review": True,
                "year": 2022,
                "marks": 3,
                "source_page": 5,
            },
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "output.xlsx")
            build_workbook(output_path, raw_records, None, [], None, extracted_texts=[])
            wb = load_workbook(output_path)

        self.assertIn("Topic Analysis", wb.sheetnames)
        sheet = wb["Topic Analysis"]
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            [
                "Unit / CO",
                "Topic (Syllabus / Concept)",
                "Frequency (Total times asked)",
                "Total Marks",
                "Primary Question Type (Theory / Derivation / Numerical)",
                "Importance Level (Critical / High / Medium / Low)",
                "2021-22 (Questions)",
                "2022-23 (Questions)",
            ],
        )

    def test_topic_analysis_aggregates_and_sorts_required_fields(self):
        raw_records = [
            {"question_number": "3(c')", "topic": "Unit II - Probability", "question_type": "Calculation", "year": "2022-23", "marks": 8},
            {"question_number": "3(b')", "topic": "Unit II - Probability", "question_type": "Derivation", "year": 2022, "marks": 5},
            {"question_number": "1(a)", "topic": "Unit I - Matrices", "question_type": "Essay", "year": 2021, "marks": 10},
            {"question_number": "2(a)", "topic": "Unit I - Matrices", "question_type": "Essay", "year": 2022, "marks": 10},
            {"question_number": "4", "topic": "Unit I - Vectors", "question_type": "Short Answer", "year": 2022, "marks": 2},
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "output.xlsx")
            build_workbook(output_path, raw_records, None, [], None)
            wb = load_workbook(output_path, data_only=False)
            sheet = wb["Topic Analysis"]
            rows = list(sheet.iter_rows(min_row=2, max_row=4, values_only=True))

        self.assertEqual(rows[0][:6], ("Unit I - CO1", "Matrices", 2, 20, "Theoretical", "Critical"))
        self.assertEqual(rows[0][6:], ("1(a) - 10M (T)", "2(a) - 10M (T)"))
        self.assertEqual(rows[1][:2], ("Unit I - CO1", "Vectors"))
        self.assertEqual(rows[1][5], "Low")
        self.assertEqual(rows[2][:6], ("Unit II - CO2", "Probability", 2, 13, "Numerical", "High"))
        self.assertEqual(rows[2][6], "Not Asked")
        self.assertEqual(rows[2][7], "3(c') - 8M (N), 3(b') - 5M (D)")


if __name__ == "__main__":
    unittest.main()
