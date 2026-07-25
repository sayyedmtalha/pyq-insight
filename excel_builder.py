from collections import Counter, OrderedDict
from pathlib import Path
from typing import List, Dict, Any
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill(fill_type="solid", fgColor="000000")
UNIT_FILLS = [
    PatternFill(fill_type="solid", fgColor="FFFFFF"),
    PatternFill(fill_type="solid", fgColor="E7E6E6"),
    PatternFill(fill_type="solid", fgColor="D9D9D9"),
]
WHITE_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
THIN_BLACK = Side(style="thin", color="000000")
TABLE_BORDER = Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK)


IMPORTANCE_ORDER = {
    "Critical": 0,
    "High": 1,
    "Medium": 2,
    "Low": 3,
}


def _split_unit_and_topic(record: Dict[str, Any]) -> tuple[str, str]:
    """Return a presentation-ready unit and syllabus topic without duplicating the unit."""
    raw_topic = str(record.get("topic") or "Unclassified").strip()
    raw_subtopic = str(record.get("subtopic") or "").strip()
    explicit_unit = str(record.get("unit") or record.get("module") or "").strip()

    unit_match = re.search(
        r"^\s*((?:unit|module)\s*(?:[ivxlcdm]+|\d+))\s*[-–—:.)]*\s*",
        f"{explicit_unit} {raw_topic}",
        re.IGNORECASE,
    )
    if unit_match:
        raw_unit = unit_match.group(1)
    else:
        raw_unit = explicit_unit or "Unit 1"

    unit_str = re.sub(r"^(?:unit|module)\s*", "", raw_unit, flags=re.IGNORECASE).strip().upper()
    roman_map = {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V", "6": "VI"}
    roman = roman_map.get(unit_str, unit_str if unit_str in {"I", "II", "III", "IV", "V", "VI"} else "I")
    co_num = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6}.get(roman, int(unit_str) if unit_str.isdigit() else 1)
    unit = f"Unit {roman} - CO{co_num}"

    topic = raw_subtopic or re.sub(r"^\s*(?:unit|module)\s*(?:[ivxlcdm]+|\d+)\s*[-–—:.)]*\s*", "", raw_topic, flags=re.IGNORECASE).strip() or raw_topic
    return unit, topic


def _academic_year(value: Any) -> tuple[int, str] | None:
    """Normalize a paper year to the academic-year label required by the output."""
    if value is None or value == "":
        return None
    match = re.search(r"(?:19|20)\d{2}", str(value))
    if not match:
        return None
    start = int(match.group(0))
    return start, f"{start}-{str(start + 1)[-2:]}"


def _importance(frequency: int, total_marks: float, years: List[int], all_years: List[int]) -> str:
    """Rank topics cleanly into Critical, High, Medium, or Low."""
    if frequency >= 4 or total_marks >= 20:
        return "Critical"
    if frequency >= 3 or total_marks >= 12:
        return "High"
    if frequency >= 2 or total_marks >= 6:
        return "Medium"
    return "Low"


def _short_type(question_type: Any) -> str:
    normalized = _normalized_type(question_type)
    mapping = {"Numerical": "N", "Theoretical": "T", "Derivation": "D"}
    return mapping.get(normalized, normalized[:1].upper() or "?")


def _normalized_type(question_type: Any) -> str:
    value = str(question_type or "").strip().lower()
    if any(token in value for token in ("deriv", "proof")):
        return "Derivation"
    if any(token in value for token in ("calculation", "numerical", "problem", "compute", "quantitative")):
        return "Numerical"
    if any(token in value for token in ("essay", "short", "long", "theory", "theoretical", "mcq", "objective", "answer")):
        return "Theoretical"
    return "Theoretical" if not value or value == "unknown" else str(question_type).strip()


def _summary_type(records: List[Dict[str, Any]]) -> str:
    types = {_normalized_type(record.get("question_type")) for record in records}
    if types == {"Theoretical", "Derivation"}:
        return "Theoretical / Derivation"
    if types == {"Theoretical", "Numerical"}:
        return "Theoretical / Numerical"
    if len(types) == 1:
        return next(iter(types))
    return Counter(_normalized_type(record.get("question_type")) for record in records).most_common(1)[0][0]


def _question_label(record: Dict[str, Any]) -> str:
    number = record.get("question_number", "?")
    marks = record.get("marks")
    question_type = _short_type(record.get("question_type"))
    if isinstance(marks, (int, float)) and marks > 0:
        mark_text = f"{marks:g}M"
        return f"{number} - {mark_text} ({question_type})"
    return f"{number} ({question_type})"


def _apply_table_style(ws, start_row: int, end_row: int, end_col: int, name: str) -> None:
    if end_row <= start_row:
        return
    ref = f"A{start_row}:{ws.cell(end_row, end_col).coordinate}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    ws.add_table(table)


def _format_sheet(ws, freeze_rows: int = 1) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{freeze_rows + 1}"
    for row in ws.iter_rows():
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_topic_analysis(ws, raw_records: List[Dict[str, Any]], taxonomy: Any = None) -> None:
    year_lookup = {_academic_year(record.get("year")) for record in raw_records}
    year_lookup.discard(None)
    years = sorted(year_lookup, key=lambda item: item[0])
    headers = ["Unit / CO", "Topic (Syllabus / Concept)", "Frequency (Total times asked)", "Total Marks", "Primary Question Type (Theory / Derivation / Numerical)", "Importance Level (Critical / High / Medium / Low)"]
    headers.extend([f"{label} (Questions)" for _, label in years])

    ws.title = "Topic Analysis"
    ws.append(headers)

    grouped: OrderedDict[tuple[str, str], List[Dict[str, Any]]] = OrderedDict()
    unit_sequence: Dict[str, int] = {}
    subtopic_sequence: Dict[tuple[str, str], int] = {}

    # 1. Initialize declared syllabus subtopics in exact Unit order
    if taxonomy is not None and getattr(taxonomy, "subtopics", None):
        for unit_topic, subtopics in taxonomy.subtopics.items():
            unit, _ = _split_unit_and_topic({"topic": unit_topic, "unit": unit_topic})
            unit_key = unit.lower()
            unit_sequence.setdefault(unit_key, len(unit_sequence))
            for sub_idx, subtopic in enumerate(subtopics):
                sub_name = str(subtopic).strip()
                if sub_name:
                    grouped.setdefault((unit, sub_name), [])
                    subtopic_sequence[(unit_key, sub_name.lower())] = sub_idx

    # 2. Map question records to their declared (Unit, Subtopic) bucket
    for record in raw_records:
        unit, topic = _split_unit_and_topic(record)
        matched_key = (unit, topic)
        for (dec_unit, dec_topic) in list(grouped.keys()):
            if dec_unit == unit and dec_topic.lower() == topic.lower():
                matched_key = (dec_unit, dec_topic)
                break
        grouped.setdefault(matched_key, []).append(record)

    topic_rows = []
    available_years = [year for year, _ in years]
    for (unit, topic), records in grouped.items():
        marks = sum(float(record.get("marks") or 0) for record in records)
        record_years = sorted({item[0] for record in records if (item := _academic_year(record.get("year")))})
        total_marks = int(round(marks))
        row = [
            unit,
            topic,
            len(records),
            total_marks,
            _summary_type(records) if records else "Theoretical",
            _importance(len(records), marks, record_years, available_years),
        ]
        for year, _ in years:
            items = [_question_label(record) for record in records if (item := _academic_year(record.get("year"))) and item[0] == year]
            row.append(", ".join(items) if items else "Not Asked")
        topic_rows.append(row)

    def topic_sort_key(row: List[Any]) -> tuple:
        unit_key = str(row[0]).lower()
        topic_key = str(row[1]).lower()
        if unit_key in unit_sequence:
            return (0, unit_sequence[unit_key], subtopic_sequence.get((unit_key, topic_key), 999), IMPORTANCE_ORDER.get(row[5], 99), topic_key)
        return (1, unit_key, 0, IMPORTANCE_ORDER.get(row[5], 99), topic_key)

    topic_rows.sort(key=topic_sort_key)
    for row in topic_rows:
        ws.append(row)

    data_last_row = ws.max_row

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER
    unit_fill_by_name: Dict[str, PatternFill] = {}
    for row in ws.iter_rows(min_row=2, max_row=data_last_row):
        unit = str(row[0].value or "Unassigned")
        if unit not in unit_fill_by_name:
            unit_fill_by_name[unit] = UNIT_FILLS[len(unit_fill_by_name) % len(UNIT_FILLS)]
        row_fill = unit_fill_by_name[unit]
        for cell in row:
            cell.fill = row_fill
            cell.border = TABLE_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row[1].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.auto_filter.ref = f"A1:{ws.cell(data_last_row, len(headers)).coordinate}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    widths = [14, 44, 13, 14, 19, 15] + [29] * (len(headers) - 6)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width
    for row_index in range(2, data_last_row + 1):
        ws.row_dimensions[row_index].height = 27
    ws.column_dimensions["D"].number_format = "#,##0"
    ws.conditional_formatting.add(f"C2:C{data_last_row}", CellIsRule(operator="greaterThanOrEqual", formula=["3"], fill=PatternFill(fill_type="solid", fgColor="8BE28B")))

def _build_raw_data(ws, raw_records: List[Dict[str, Any]]) -> None:
    headers = ["question_number", "text", "topic", "subtopic", "difficulty", "question_type", "marks", "confidence", "manual_review", "year", "source_page"]
    ws.append(headers)
    for record in raw_records:
        ws.append([
            record.get("question_number", ""), record.get("text", ""), record.get("topic", ""), record.get("subtopic", ""),
            record.get("difficulty", ""), record.get("question_type", ""), record.get("marks", ""), record.get("confidence", ""),
            "Yes" if record.get("manual_review") else "No", record.get("year", ""), record.get("source_page", ""),
        ])
    _format_sheet(ws)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
    ws.auto_filter.ref = f"A1:{ws.cell(ws.max_row, len(headers)).coordinate}"
    widths = [16, 70, 26, 26, 14, 16, 12, 12, 16, 12, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width
    ws.column_dimensions["G"].number_format = "#,##0.0"
    ws.column_dimensions["H"].number_format = "0%"
    _apply_table_style(ws, 1, ws.max_row, len(headers), "RawQuestionTable")


def build_workbook(output_path: str, raw_records: List[Dict[str, Any]], topic_pivot: pd.DataFrame, gaps: List[Dict[str, Any]], trends: pd.DataFrame, extracted_texts: List[Dict[str, Any]] | None = None, taxonomy: Any = None) -> str:
    wb = Workbook()
    ws_topic = wb.active
    _build_topic_analysis(ws_topic, raw_records, taxonomy=taxonomy)

    output = Path(output_path)
    wb.save(output)
    return str(output)
